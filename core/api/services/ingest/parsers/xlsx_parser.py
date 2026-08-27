"""XLSX parser for universal ingestion."""
from __future__ import annotations

import hashlib
import os
from pathlib import Path
from typing import Any, BinaryIO

from openpyxl import load_workbook

from core.api.services.ingest.serializers.xlsx_to_markdown import (
    MAX_ROWS_PER_SHEET,
    serialize_sheet_to_markdown,
)
from core.api.services.ingest.xlsx_privacy import (
    PROPRIETARY_DETECTOR,
    neutral_xlsx_summary,
    neutral_xlsx_title,
    xlsx_sheet_names,
)

STREAMING_THRESHOLD_BYTES = 10 * 1024 * 1024
PARSER_USED = "openpyxl"


def _stream_sha256(source: BinaryIO) -> str:
    digest = hashlib.sha256()
    position = source.tell()
    try:
        source.seek(0)
        for chunk in iter(lambda: source.read(65536), b""):
            digest.update(chunk)
    finally:
        source.seek(position)
    return digest.hexdigest()


def _source_identity(stat_result: os.stat_result) -> tuple[int, int, int, int]:
    return (
        int(stat_result.st_dev),
        int(stat_result.st_ino),
        int(stat_result.st_size),
        int(stat_result.st_mtime_ns),
    )


def _validate_source_stable(
    path: Path,
    source: BinaryIO,
    *,
    identity: tuple[int, int, int, int],
    sha256: str,
) -> None:
    try:
        descriptor_identity = _source_identity(os.fstat(source.fileno()))
        path_identity = _source_identity(path.stat(follow_symlinks=False))
    except OSError:
        raise ValueError("XLSX source changed during parse") from None
    if (
        descriptor_identity != identity
        or path_identity != identity
        or _stream_sha256(source) != sha256
    ):
        raise ValueError("XLSX source changed during parse")


def _workbook_markdown(title: str, sheets: list[dict[str, Any]]) -> str:
    parts = [f"# {title}"]
    for sheet in sheets:
        parts.extend(
            [
                "",
                f"## Sheet: {sheet['name']}",
                "",
                sheet["markdown"],
            ]
        )
    return "\n".join(parts).strip() + "\n"


def parse_xlsx(path: Path) -> dict[str, Any]:
    """Parse an `.xlsx` workbook into markdown text and sheet metadata.

    Security contract:
    - `.xlsm` is rejected before openpyxl touches the file.
    - Formulae are never evaluated (`data_only=True` reads cached values).
    - External workbook links are disabled (`keep_links=False`).
    - Files over 10MB use openpyxl's read-only streaming mode.
    """
    suffix = path.suffix.lower()
    if suffix == ".xlsm":
        raise ValueError("XLSM macro-enabled workbooks are not supported")
    if suffix != ".xlsx":
        raise ValueError(f"Unsupported spreadsheet extension: {suffix or '<none>'}")

    nofollow = getattr(os, "O_NOFOLLOW", 0)
    fd = os.open(path, os.O_RDONLY | nofollow)
    with os.fdopen(fd, "rb") as source:
        source_stat = os.fstat(source.fileno())
        identity = _source_identity(source_stat)
        size = int(source_stat.st_size)
        sha256 = _stream_sha256(source)
        source.seek(0)
        manifest_sheet_count = len(xlsx_sheet_names(source))
        source.seek(0)
        read_only = size > STREAMING_THRESHOLD_BYTES
        workbook = load_workbook(
            filename=source,
            read_only=read_only,
            data_only=True,
            keep_links=False,
        )
        try:
            sheets: list[dict[str, Any]] = []
            proprietary = manifest_sheet_count > 1
            if not proprietary:
                for index, worksheet in enumerate(workbook.worksheets, start=1):
                    row_count = worksheet.max_row or 0
                    sheets.append(
                        {
                            "name": worksheet.title,
                            "index": index,
                            "row_count": row_count,
                            "column_count": worksheet.max_column,
                            "markdown": serialize_sheet_to_markdown(worksheet),
                            "truncated": row_count > MAX_ROWS_PER_SHEET,
                        }
                    )

            structure = {
                "kind": "xlsx",
                "bytes": size,
                "sha256": sha256,
                "streaming": read_only,
                "data_only": True,
                "keep_links": False,
                "max_rows_per_sheet": MAX_ROWS_PER_SHEET,
                "proprietary": proprietary,
                "proprietary_detector": PROPRIETARY_DETECTOR,
            }
            if not proprietary:
                structure["sheet_count"] = manifest_sheet_count
                structure["sheets"] = [
                    {
                        key: value
                        for key, value in sheet.items()
                        if key != "markdown"
                    }
                    for sheet in sheets
                ]
            title = neutral_xlsx_title(sha256)
            text = (
                neutral_xlsx_summary(
                    sha256,
                    sheet_count=manifest_sheet_count,
                )
                if proprietary
                else _workbook_markdown(title, sheets)
            )
            result = {
                "frontmatter": {
                    "type": "file",
                    "title": title,
                    "tags": ["xlsx", "spreadsheet"],
                },
                "text": text,
                "structure": structure,
                "parser_used": PARSER_USED,
            }
        finally:
            workbook.close()
        _validate_source_stable(
            path,
            source,
            identity=identity,
            sha256=sha256,
        )
        return result
